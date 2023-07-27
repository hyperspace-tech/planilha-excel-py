#Exercicio proposto pelo professor do canal do "Dev Aprender | Jhonatan de Souza"

import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('meus computadores')

computadores_page = book['meus computadores']
computadores_page.append(['eletronica', 'memoria ram', 'preço'])
computadores_page.append(['computador 1', '8gb ram', 'R$2500'])
computadores_page.append(['computador 2', '16 ram', 'R$5500'])
computadores_page.append(['computador 3', '32 ram', 'R$8500'])

book.save('Meus computadores.xlsx')